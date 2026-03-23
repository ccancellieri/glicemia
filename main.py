import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib.ticker import MaxNLocator
import numpy as np
import os
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.chart import (
    LineChart,
    BarChart,
    AreaChart,
    Reference,
    Series,
)
from openpyxl.chart.axis import DateAxis, TextAxis
import io
import warnings
from datetime import datetime, timedelta, time

# Ignora avvisi comuni di matplotlib per un output più pulito
warnings.filterwarnings("ignore", category=UserWarning)

# --- COSTANTI DI CONFIGURAZIONE ---
TARGET_RANGE_LOW = 70
TARGET_RANGE_HIGH = 180
HYPO_THRESHOLD = 70
HYPER_THRESHOLD = 180
VERY_HIGH_THRESHOLD = 250
DEFAULT_INSULIN_ACTION_HOURS = 3.0 # Valore di default per la prima generazione

def parse_and_clean_data(file_path):
    """
    Parsa un report CareLink .csv, individua la tabella dati principale,
    la pulisce e la prepara per l'analisi. Ora include la colonna Suspend.
    """
    print(f"1. Inizio parsing e pulizia del file: {os.path.basename(file_path)}")
    try:
        with open(file_path, 'r', encoding='latin1') as f:
            lines = f.readlines()
    except FileNotFoundError:
        print(f"ERRORE: File non trovato: {file_path}")
        return None, {}

    header_row_index = -1
    header_keywords = ['Index', 'Date', 'Time', 'BG Reading', 'Bolus Volume Delivered', 'Suspend']
    for i, line in enumerate(lines):
        if all(keyword in line for keyword in header_keywords):
            header_row_index = i
            break
            
    if header_row_index == -1:
        print("ATTENZIONE: Intestazione della tabella dati principale non trovata. Impossibile processare il file.")
        return None, {}

    csv_buffer = io.StringIO(''.join(lines[header_row_index:]))
    df = pd.read_csv(csv_buffer, sep=';', on_bad_lines='skip', low_memory=False)

    df.columns = df.columns.str.strip()
    df.rename(columns={
        'Sensor Glucose (mg/dL)': 'Glicemia',
        'BWZ Carb Input (grams)': 'Carboidrati',
        'Bolus Volume Delivered (U)': 'Bolo',
        'Basal Rate (U/h)': 'Basale',
        'BWZ Insulin Sensitivity (mg/dL/U)': 'ISF',
        'BWZ Carb Ratio (g/U)': 'IC',
        'Suspend': 'Sospensione'
    }, inplace=True)

    df['Date'] = pd.to_datetime(df['Date'], format='%Y/%m/%d', errors='coerce')
    df['Time'] = pd.to_datetime(df['Time'], format='%H:%M:%S', errors='coerce').dt.time
    df.dropna(subset=['Date', 'Time'], inplace=True)
    df['Timestamp'] = df.apply(lambda r: datetime.combine(r['Date'].date(), r['Time']), axis=1)
    df = df.set_index('Timestamp').sort_index()

    cols_to_numeric = ['Glicemia', 'Carboidrati', 'Bolo', 'Basale', 'ISF', 'IC']
    for col in cols_to_numeric:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col].astype(str).str.replace(',', '.'), errors='coerce')

    # CORREZIONE: Gestione robusta della colonna 'Sospensione' per evitare TypeError
    if 'Sospensione' in df.columns:
        df['Sospensione'] = pd.to_numeric(df['Sospensione'], errors='coerce').fillna(0).astype(int)

    df['Bolo_Correzione'] = df['Bolo'].where((df['Carboidrati'].fillna(0) == 0) & (df['Bolo'] > 0), 0)
    
    df = df[['Glicemia', 'Carboidrati', 'Bolo', 'Basale', 'Bolo_Correzione', 'ISF', 'IC', 'Sospensione']].copy()
    df = df[~df.index.duplicated(keep='last')]
    df['Glicemia'] = df['Glicemia'].interpolate(method='time', limit_direction='both')
    df.dropna(subset=['Glicemia'], inplace=True)
    
    patient_info = {}
    try:
        patient_info['name'] = os.getenv("PATIENT_NAME", "Patient")
        patient_info['start_date'] = df.index.min().strftime('%d/%m/%Y')
        patient_info['end_date'] = df.index.max().strftime('%d/%m/%Y')
    except (IndexError, AttributeError):
        patient_info = {'name': 'N/D', 'start_date': 'N/D', 'end_date': 'N/D'}

    print("2. Parsing e pulizia completati.")
    return df, patient_info

def prepare_simulation_data(df):
    if df.empty: return pd.DataFrame()
    print("   - Preparazione dati per i simulatori...")
    df_resampled = df.resample('5min').last()
    df_resampled['Bolo'] = df['Bolo'].resample('5min').sum()
    df_resampled['Basale_Rate'] = df['Basale'].ffill()
    df_resampled['Glicemia'] = df_resampled['Glicemia'].interpolate(method='time').round(0)
    df_resampled['Basal_Input'] = (df_resampled['Basale_Rate'].fillna(0) / 12)
    df_resampled['Bolus_Input'] = df_resampled['Bolo'].fillna(0)
    df_resampled['Total_Insulin_Input'] = df_resampled['Basal_Input'] + df_resampled['Bolus_Input']
    df_resampled['Glicemia_Ipo'] = np.where(df_resampled['Glicemia'] < HYPO_THRESHOLD, df_resampled['Glicemia'], np.nan)
    df_resampled['Glicemia_Target'] = np.where((df_resampled['Glicemia'] >= TARGET_RANGE_LOW) & (df_resampled['Glicemia'] <= TARGET_RANGE_HIGH), df_resampled['Glicemia'], np.nan)
    df_resampled['Glicemia_Iper'] = np.where(df_resampled['Glicemia'] > HYPER_THRESHOLD, df_resampled['Glicemia'], np.nan)
    return df_resampled[['Glicemia_Ipo', 'Glicemia_Target', 'Glicemia_Iper', 'Basal_Input', 'Bolus_Input', 'Total_Insulin_Input']].reset_index()

def create_weekly_model(df):
    if df.empty: return pd.DataFrame()
    print("   - Creazione modello settimanale medio...")
    df_resampled = df.resample('5min').sum()
    df_resampled['Time'] = df_resampled.index.time
    average_day = df_resampled.groupby('Time')[['Carboidrati', 'Bolo_Correzione']].mean()
    start_of_week = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0) - timedelta(days=datetime.now().weekday())
    week_index = pd.date_range(start=start_of_week, periods=7 * 288, freq='5min')
    weekly_model = pd.DataFrame(index=week_index)
    weekly_model['Time'] = weekly_model.index.time
    weekly_model = pd.merge(weekly_model, average_day, on='Time', how='left').set_index(weekly_model.index)
    return weekly_model[['Carboidrati', 'Bolo_Correzione']].reset_index().rename(columns={'index': 'Timestamp'})

def generate_static_visualizations(df, output_folder):
    if df.empty: return {}
    print("3. Generazione grafici di analisi statica...")
    paths = {}
    plt.style.use('seaborn-v0_8-whitegrid')
    in_range = ((df['Glicemia'] >= TARGET_RANGE_LOW) & (df['Glicemia'] <= TARGET_RANGE_HIGH)).sum()
    hypo = (df['Glicemia'] < HYPO_THRESHOLD).sum()
    hyper = (df['Glicemia'] > HYPER_THRESHOLD).sum()
    total = in_range + hypo + hyper or 1
    labels, sizes, colors = [f'In Target ({TARGET_RANGE_LOW}-{TARGET_RANGE_HIGH})', f'Ipo (< {HYPO_THRESHOLD})', f'Iper (> {HYPER_THRESHOLD})'], [in_range/total, hypo/total, hyper/total], ['#4CAF50', '#2196F3', '#FFC107']
    fig1, ax1 = plt.subplots(figsize=(10, 6))
    wedges, texts, autotexts = ax1.pie(sizes, colors=colors, autopct='%1.1f%%', startangle=90, pctdistance=0.85, wedgeprops=dict(width=0.4))
    plt.setp(autotexts, size=10, weight="bold", color="white")
    ax1.axis('equal'); ax1.set_title('Tempo negli Intervalli (TIR)'); ax1.legend(wedges, labels, title="Legenda", loc="center left", bbox_to_anchor=(1, 0, 0.5, 1))
    path = os.path.join(output_folder, 'tir_chart.png'); plt.savefig(path, bbox_inches='tight'); plt.close(fig1); paths['tir'] = path
    print("4. Grafici statici generati.")
    return paths

def analyze_suspensions(df, sim_df_full):
    print("5. Analisi eventi di sospensione pompa...")
    if 'Sospensione' not in df.columns or df['Sospensione'].sum() == 0:
        print("   - Nessun evento di sospensione esplicito trovato nei dati.")
        return pd.DataFrame()

    suspension_events = df[df['Sospensione'] == 1]
    if suspension_events.empty:
        return pd.DataFrame()
    
    analysis = []
    sim_df_indexed = sim_df_full.set_index('Timestamp')
    
    for timestamp, row in suspension_events.iterrows():
        start_window = timestamp - timedelta(hours=2)
        window_df = sim_df_indexed.loc[start_window:timestamp]
        if window_df.empty: continue
        
        window_df['Glicemia_Unita'] = window_df['Glicemia_Ipo'].fillna(window_df['Glicemia_Target']).fillna(window_df['Glicemia_Iper'])
        if window_df['Glicemia_Unita'].dropna().empty: continue

        start_glucose = window_df['Glicemia_Unita'].dropna().iloc[0]
        end_glucose = window_df['Glicemia_Unita'].dropna().iloc[-1]
        glucose_drop = start_glucose - end_glucose
        
        intervals = DEFAULT_INSULIN_ACTION_HOURS * 12
        iob = 0
        if intervals > 0:
            iob_series = window_df['Total_Insulin_Input'].ewm(span=intervals, adjust=False).mean() * intervals
            iob = iob_series.iloc[-1] if not iob_series.empty else 0

        analysis.append({
            'Data e Ora Sospensione': timestamp,
            'Glicemia Inizio Finestra (mg/dL)': int(start_glucose),
            'Glicemia a Sospensione (mg/dL)': int(end_glucose),
            'Calo Glicemico (mg/dL/2h)': int(glucose_drop),
            'IOB Stimato a Sospensione (U)': f"{iob:.2f}",
            'Boli nelle 2h prec. (U)': window_df['Bolus_Input'].sum()
        })
    return pd.DataFrame(analysis)

def analyze_hypo_episodes(df):
    print("6. Analisi dettagliata episodi di ipoglicemia...")
    hypo_episodes = []
    in_hypo = False
    episode_start, current_episode_data, previous_timestamp = None, [], None
    hypo_data = df[df['Glicemia'] < HYPO_THRESHOLD]
    if hypo_data.empty: return pd.DataFrame()
    for timestamp, row in hypo_data.iterrows():
        if not in_hypo:
            in_hypo, episode_start, current_episode_data = True, timestamp, [row['Glicemia']]
        else:
            if previous_timestamp and (timestamp - previous_timestamp) < timedelta(minutes=15):
                current_episode_data.append(row['Glicemia'])
            else:
                if previous_timestamp:
                    duration = (previous_timestamp - episode_start).total_seconds() / 60
                    nadir = min(current_episode_data)
                    context_window = df.loc[episode_start - timedelta(hours=3): episode_start]
                    recent_bolo = context_window[context_window['Bolo'] > 0]
                    context_str = f"Bolo di {recent_bolo['Bolo'].sum():.1f} U nelle 3h prec." if not recent_bolo.empty else "Nessun bolo recente."
                    hypo_episodes.append({'Inizio': episode_start, 'Durata (min)': int(duration), 'Nadir (mg/dL)': int(nadir), 'Contesto': context_str})
                episode_start, current_episode_data = timestamp, [row['Glicemia']]
        previous_timestamp = timestamp
    if in_hypo and episode_start and previous_timestamp:
        duration = (previous_timestamp - episode_start).total_seconds() / 60
        nadir = min(current_episode_data)
        context_window = df.loc[episode_start - timedelta(hours=3): episode_start]
        recent_bolo = context_window[context_window['Bolo'] > 0]
        context_str = f"Bolo di {recent_bolo['Bolo'].sum():.1f} U nelle 3h prec." if not recent_bolo.empty else "Nessun bolo recente."
        hypo_episodes.append({'Inizio': episode_start, 'Durata (min)': int(duration), 'Nadir (mg/dL)': int(nadir), 'Contesto': context_str})
    return pd.DataFrame(hypo_episodes)

def analyze_patterns_and_suggest(df):
    if df.empty: return pd.DataFrame()
    print("7. Analisi dei pattern e generazione suggerimenti...")
    analysis = []
    df_hourly = df.copy()
    df_hourly['Ora'], df_hourly['Giorno'] = df_hourly.index.hour, df_hourly.index.date
    time_slots = {'Notte (22-08)': (time(22, 0), time(7, 59)), 'Giorno (08-16)': (time(8, 0), time(15, 59)), 'Sera (16-22)': (time(16, 0), time(21, 59))}
    total_days = len(df_hourly['Giorno'].unique()) or 1
    for name, (start, end) in time_slots.items():
        try:
            slot_df = df_hourly.between_time(start, end)
        except ValueError: 
            slot_df = df_hourly[(df_hourly.index.time >= start) | (df_hourly.index.time <= end)]
            
        if slot_df.empty: continue
        hypo_days = slot_df[slot_df['Glicemia'] < HYPO_THRESHOLD]['Giorno'].nunique()
        if hypo_days / total_days > 0.3:
            avg_hypo = slot_df[slot_df['Glicemia'] < HYPO_THRESHOLD]['Glicemia'].mean()
            analysis.append({'Fascia Oraria': name, 'Problema': 'IPOGLICEMIA Ricorrente', 'Dettagli': f"Nel {hypo_days/total_days:.0%} dei giorni. Media ipo: {avg_hypo:.0f} mg/dL.", 'Azione Suggerita': "1. Considerare Obiettivo Glicemico più alto.\n2. Verificare se TIA è troppo lungo."})
        hyper_days = slot_df[slot_df['Glicemia'] > HYPER_THRESHOLD]['Giorno'].nunique()
        if hyper_days / total_days > 0.5:
            avg_hyper = slot_df[slot_df['Glicemia'] > HYPER_THRESHOLD]['Glicemia'].mean()
            analysis.append({'Fascia Oraria': name, 'Problema': 'IPERGLICEMIA Ricorrente', 'Dettagli': f"Nel {hyper_days/total_days:.0%} dei giorni. Media iper: {avg_hyper:.0f} mg/dL.", 'Azione Suggerita': "1. Valutare se rapporto I:C è corretto.\n2. Verificare se ISF è adeguato."})
    gmi = (3.31 + 0.02392 * df['Glicemia'].mean())
    tir_perc = (df['Glicemia'].between(TARGET_RANGE_LOW, TARGET_RANGE_HIGH).sum() / len(df)) * 100
    summary = {'Fascia Oraria': 'Generale', 'Problema': 'Riepilogo', 'Dettagli': f"Glucosio Medio: {df['Glicemia'].mean():.0f} mg/dL\nGMI: {gmi:.1f}%\nTempo in Target: {tir_perc:.1f}%", 'Azione Suggerita': "Usare i dati delle fasce orarie per un 'fine-tuning' mirato."}
    analysis.insert(0, summary)
    return pd.DataFrame(analysis)

def create_interactive_excel_report(patient_info, df, sim_df, weekly_model_df, analysis_df, hypo_analysis_df, suspension_df, image_paths, output_file):
    print(f"8. Creazione del report Excel INTERATTIVO: {output_file}")
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        wb = writer.book
        title_font = Font(size=18, bold=True, color="0D47A1")
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="42A5F5", end_color="42A5F5", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        def style_sheet(ws, df_cols):
            for i, col in enumerate(ws.columns):
                max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
                ws.column_dimensions[get_column_letter(i+1)].width = min(max_length + 2, 50)
            start_row = 5 if ws.title not in ["Dati Dettagliati", "Dati_Simulazione_Reale", "Dati_Modello_Settimanale"] else 1
            for cell in ws[start_row]: cell.font, cell.fill = header_font, header_fill

        def set_chart_axis_style(chart):
            chart.x_axis.number_format = 'd-mmm h:mm'
            chart.x_axis.tickLblSkip = 12 
            chart.x_axis.scaling.orientation = "maxMin"
        
        # --- Fogli di Analisi Statica ---
        analysis_df.to_excel(writer, sheet_name="Dashboard & Suggerimenti", index=False, startrow=4)
        ws1 = writer.sheets["Dashboard & Suggerimenti"]; ws1['A1'] = f"Report Analisi Diabete - {patient_info.get('name', 'N/D')}"
        ws1['A1'].font, ws1['A2'] = title_font, f"Periodo: {patient_info.get('start_date', 'N/D')} - {patient_info.get('end_date', 'N/D')}"
        if 'tir' in image_paths: ws1.add_image(Image(image_paths['tir']), 'A15')
        style_sheet(ws1, analysis_df.columns)

        hypo_analysis_df.to_excel(writer, sheet_name="Analisi Ipoglicemie", index=False, startrow=4)
        ws2 = writer.sheets["Analisi Ipoglicemie"]; ws2['A1'] = "Analisi Dettagliata Episodi di Ipoglicemia"; ws2['A1'].font = title_font
        if hypo_analysis_df.empty: ws2['A6'] = "Nessun evento di ipoglicemia (< 70 mg/dL) è stato registrato in questo periodo."
        style_sheet(ws2, hypo_analysis_df.columns)

        suspension_df.to_excel(writer, sheet_name="Analisi Sospensioni Pompa", index=False, startrow=4)
        ws3 = writer.sheets["Analisi Sospensioni Pompa"]; ws3['A1'] = "Analisi Eventi di Sospensione della Pompa"; ws3['A1'].font = title_font
        if suspension_df.empty: ws3['A6'] = "Nessun evento di sospensione esplicito (valore '1' nella colonna Suspend) è stato trovato nei dati."
        style_sheet(ws3, suspension_df.columns)
        
        # --- Fogli Dati per Simulatori (VISIBILI) ---
        sim_df.to_excel(writer, sheet_name="Dati_Simulazione_Reale", index=False)
        style_sheet(writer.sheets["Dati_Simulazione_Reale"], sim_df.columns)
        weekly_model_df.to_excel(writer, sheet_name="Dati_Modello_Settimanale", index=False)
        style_sheet(writer.sheets["Dati_Modello_Settimanale"], weekly_model_df.columns)

        # --- Pannello di Controllo e Simulatore Parametri ---
        ws_param = wb.create_sheet("Simulatore Parametri")
        ws_param['A1'] = "Simulatore Predittivo Settimanale"; ws_param['A1'].font = title_font
        ws_param['A2'] = "Modifichi i parametri nelle celle gialle per simulare l'impatto sull'IOB medio settimanale."
        ws_param['B4'] = "PARAMETRI DI SIMULAZIONE"; ws_param['B4'].font = Font(bold=True)
        params_header = ["Fascia Oraria", "Rapporto I:C (g/U)", "Sensibilità (mg/dL/U)", "Basale (U/ora)"]
        fasce = ["Notte (22-08)", "Giorno (08-16)", "Sera (16-22)"]
        ws_param.append(params_header); 
        for i, fascia in enumerate(fasce): ws_param.append([fascia, 10, 50, 0.8])
        ws_param['A10'] = "Tempo Azione Insulina (ore):"; ws_param['A10'].font = Font(bold=True)
        ws_param['B10'] = DEFAULT_INSULIN_ACTION_HOURS
        for row in ws_param['B6:D8']:
            for cell in row: cell.fill = yellow_fill
        ws_param['B10'].fill = yellow_fill
        
        ws_param['K4'] = "Tabella di Supporto VLOOKUP"; ws_param['K4'].font = Font(bold=True)
        ws_param.append([]); ws_param['K5'] = "Ora Inizio"; ws_param['L5'] = "I:C"; ws_param['M5'] = "ISF"; ws_param['N5'] = "Basale"
        ws_param['K6'] = time(0, 0); ws_param['L6'] = "=B6"; ws_param['M6'] = "=C6"; ws_param['N6'] = "=D6"
        ws_param['K7'] = time(8, 0); ws_param['L7'] = "=B7"; ws_param['M7'] = "=C7"; ws_param['N7'] = "=D7"
        ws_param['K8'] = time(16, 0); ws_param['L8'] = "=B8"; ws_param['M8'] = "=C8"; ws_param['N8'] = "=D8"
        ws_param['K9'] = time(22, 0); ws_param['L9'] = "=B6"; ws_param['M9'] = "=C6"; ws_param['N9'] = "=D6"

        ws_param.append([]); ws_param.append(["Timestamp", "IOB Simulata (U)"])
        num_rows_model = len(weekly_model_df)
        for r in range(num_rows_model):
            row_num = r + 13
            ws_param.cell(row=row_num, column=1, value=f"=Dati_Modello_Settimanale!A{r+2}")
            time_check = f"MOD(A{row_num},1)"
            ic_formula = f"VLOOKUP({time_check},$K$6:$N$9,2,TRUE)"
            basal_formula = f"VLOOKUP({time_check},$K$6:$N$9,4,TRUE)"
            basal_input = f"({basal_formula}/12)"
            food_input = f"IFERROR(Dati_Modello_Settimanale!B{r+2}/{ic_formula},0)"
            corr_input = f"Dati_Modello_Settimanale!C{r+2}"
            total_input = f"SUM({basal_input},{food_input},{corr_input})"
            intervals = f"($B$10*12)"
            if r == 0: ws_param.cell(row=row_num, column=2, value=f"={total_input}")
            else: ws_param.cell(row=row_num, column=2, value=f"=IF({intervals}<=0, {total_input}, B{row_num-1}*(1-1/{intervals})+{total_input})")
        
        chart_param = LineChart(); chart_param.title = "Simulazione IOB su Modello Settimanale Medio"; chart_param.y_axis.title = "IOB (U)"
        cats_param = Reference(ws_param, min_col=1, min_row=13, max_row=num_rows_model+12)
        data_param = Reference(ws_param, min_col=2, min_row=12, max_row=num_rows_model+12)
        chart_param.add_data(data_param, titles_from_data=True); chart_param.set_categories(cats_param)
        set_chart_axis_style(chart_param); ws_param.add_chart(chart_param, "F2"); chart_param.height, chart_param.width = 15, 40

        # --- Simulatore IOB (Dati Reali) ---
        ws_sim = wb.create_sheet("Simulatore IOB (Dati Reali)")
        ws_sim['A1'] = "Simulatore Glicemia e IOB (Dati Reali)"; ws_sim['A1'].font = title_font
        ws_sim['A2'] = "Controllato dalla cella gialla nel foglio 'Simulatore Parametri'"
        header_sim1 = ["Timestamp", "Glicemia Ipo", "Glicemia Target", "Glicemia Iper", "IOB Calcolato (U)"]
        ws_sim.append(header_sim1);
        for cell in ws_sim[4]: cell.font, cell.fill = header_font, header_fill
        num_rows_sim = len(sim_df)
        for r in range(num_rows_sim):
            row_num = r + 5
            ws_sim.cell(row=row_num, column=1, value=f"=Dati_Simulazione_Reale!A{r+2}")
            ws_sim.cell(row=row_num, column=2, value=f"=Dati_Simulazione_Reale!B{r+2}")
            ws_sim.cell(row=row_num, column=3, value=f"=Dati_Simulazione_Reale!C{r+2}")
            ws_sim.cell(row=row_num, column=4, value=f"=Dati_Simulazione_Reale!D{r+2}")
            intervals = f"('Simulatore Parametri'!$B$10*12)"; current_input = f"Dati_Simulazione_Reale!G{r+2}"
            if r == 0: ws_sim.cell(row=row_num, column=5, value=f"={current_input}")
            else: ws_sim.cell(row=row_num, column=5, value=f"=IF({intervals}<=0, {current_input}, E{row_num-1}*(1 - 1/{intervals}) + {current_input})")
        
        area_chart = AreaChart(grouping="stacked"); area_chart.y_axis.title = "Glicemia (mg/dL)"; cats = Reference(ws_sim, min_col=1, min_row=5, max_row=num_rows_sim+4)
        series_ipo = Series(Reference(ws_sim, min_col=2, min_row=4, max_row=num_rows_sim+4), title_from_data=True); series_ipo.graphicalProperties.solidFill = "ADD8E6"
        series_target = Series(Reference(ws_sim, min_col=3, min_row=4, max_row=num_rows_sim+4), title_from_data=True); series_target.graphicalProperties.solidFill = "90EE90"
        series_iper = Series(Reference(ws_sim, min_col=4, min_row=4, max_row=num_rows_sim+4), title_from_data=True); series_iper.graphicalProperties.solidFill = "FFA07A"
        area_chart.append(series_ipo); area_chart.append(series_target); area_chart.append(series_iper); area_chart.set_categories(cats)
        line_chart = LineChart(); data_iob = Reference(ws_sim, min_col=5, min_row=4, max_row=num_rows_sim+4)
        series_iob = Series(data_iob, title_from_data=True); series_iob.graphicalProperties.line.solidFill = "FF8C00"; series_iob.graphicalProperties.line.width = 30000
        line_chart.append(series_iob); line_chart.y_axis.axId, line_chart.y_axis.title = 200, "IOB (U)"; area_chart.y_axis.crosses = "max"; area_chart += line_chart
        set_chart_axis_style(area_chart); ws_sim.add_chart(area_chart, "G5"); area_chart.height, area_chart.width = 15, 40
        
        # --- Foglio Simulatore Componenti IOB ---
        ws_comp = wb.create_sheet("Simulatore Componenti IOB")
        ws_comp['A1'] = "Simulatore Componenti IOB (Dati Reali)"; ws_comp['A1'].font = title_font
        ws_comp['A2'] = "Controllato dalla cella gialla nel foglio 'Simulatore Parametri'."
        header_sim2 = ["Timestamp", "IOB da Basale", "IOB da Boli", "IOB Totale"]
        ws_comp.append(header_sim2)
        for cell in ws_comp[4]: cell.font, cell.fill = header_font, header_fill
        for r in range(num_rows_sim):
            row_num = r + 5
            ws_comp.cell(row=row_num, column=1, value=f"=Dati_Simulazione_Reale!A{r+2}")
            intervals_comp = f"('Simulatore Parametri'!$B$10*12)"
            basal_input, bolus_input = f"Dati_Simulazione_Reale!E{r+2}", f"Dati_Simulazione_Reale!F{r+2}"
            if r == 0:
                ws_comp.cell(row=row_num, column=2, value=f"={basal_input}")
                ws_comp.cell(row=row_num, column=3, value=f"={bolus_input}")
            else:
                ws_comp.cell(row=row_num, column=2, value=f"=IF({intervals_comp}<=0, {basal_input}, B{row_num-1}*(1-1/{intervals_comp})+{basal_input})")
                ws_comp.cell(row=row_num, column=3, value=f"=IF({intervals_comp}<=0, {bolus_input}, C{row_num-1}*(1-1/{intervals_comp})+{bolus_input})")
            ws_comp.cell(row=row_num, column=4, value=f"=B{row_num}+C{row_num}")
            
        chart_comp = LineChart(); chart_comp.title = "Decomposizione IOB (Dati Reali)"; chart_comp.y_axis.title = "IOB (U)"
        cats_comp = Reference(ws_comp, min_col=1, min_row=5, max_row=num_rows_sim+4)
        chart_comp.set_categories(cats_comp)
        data_basal_iob = Reference(ws_comp, min_col=2, min_row=4, max_row=num_rows_sim+4); series_basal = Series(data_basal_iob, title_from_data=True); series_basal.graphicalProperties.line.solidFill = "8A2BE2"
        data_bolus_iob = Reference(ws_comp, min_col=3, min_row=4, max_row=num_rows_sim+4); series_bolus = Series(data_bolus_iob, title_from_data=True); series_bolus.graphicalProperties.line.solidFill = "0000FF"
        data_total_iob = Reference(ws_comp, min_col=4, min_row=4, max_row=num_rows_sim+4); series_total = Series(data_total_iob, title_from_data=True); series_total.graphicalProperties.line.solidFill = "FF8C00"; series_total.graphicalProperties.line.width = 35000
        chart_comp.append(series_basal); chart_comp.append(series_bolus); chart_comp.append(series_total)
        set_chart_axis_style(chart_comp); ws_comp.add_chart(chart_comp, "F4"); chart_comp.height, chart_comp.width = 15, 40
        
        # --- Simulatore di Correzione ---
        ws_corr = wb.create_sheet("Simulatore di Correzione")
        ws_corr['A1'] = "Simulatore Bolo di Correzione"; ws_corr['A1'].font = Font(size=18, bold=True)
        ws_corr.append([]);ws_corr.append(["PARAMETRI DI INPUT"])
        ws_corr.append(["Glicemia di Partenza (mg/dL)", 200])
        ws_corr.append(["Glicemia Obiettivo (mg/dL)", 110])
        ws_corr.append(["Fattore di Sensibilità (ISF)", 50])
        ws_corr.append(["Tempo Azione Insulina (ore)", 3])
        for row in ws_corr['B4:B7']:
            for cell in row: cell.fill = yellow_fill
        ws_corr.append([]); ws_corr.append(["RISULTATI"])
        ws_corr.append(["Bolo di Correzione Calcolato (U)", "=IFERROR(IF(B6>0,(B4-B5)/B6,0),0)"])
        ws_corr.append([]); ws_corr.append(["Minuti", "IOB Residua (U)"])
        for i in range(61):
            minutes = i * 5
            ws_corr.cell(row=i + 13, column=1, value=minutes)
            intervals = f"$B$7*12"; total_bolus = f"$B$10"
            if i == 0: ws_corr.cell(row=i + 13, column=2, value=f"={total_bolus}")
            else: ws_corr.cell(row=i + 13, column=2, value=f"=IFERROR(IF({intervals}>0,B{i+12}*(1-1/{intervals}),0),0)")
        chart_corr = LineChart(); chart_corr.title = "Curva IOB di un Bolo di Correzione"; chart_corr.y_axis.title = "IOB (U)"
        cats_corr = Reference(ws_corr, min_col=1, min_row=13, max_row=73)
        data_corr = Reference(ws_corr, min_col=2, min_row=12, max_row=73)
        chart_corr.add_data(data_corr, titles_from_data=True); chart_corr.set_categories(cats_corr)
        ws_corr.add_chart(chart_corr, "D2"); chart_corr.height, chart_corr.width = 15, 30

        # --- Foglio Dati Dettagliati ---
        df.reset_index().to_excel(writer, sheet_name="Dati Dettagliati", index=False)
        style_sheet(writer.sheets["Dati Dettagliati"], df.columns)
        
    print(f"9. Report Excel INTERATTIVO salvato con successo: {output_file}")


def main():
    files_to_analyze = [f for f in os.listdir('.') if f.lower().endswith('.csv')]
    if not files_to_analyze:
        print("\nATTENZIONE: Nessun file .csv trovato nella cartella.")
        return
    output_folder = "output_report"
    if not os.path.exists(output_folder): os.makedirs(output_folder)
    for file in files_to_analyze:
        print("-" * 50)
        clean_df, patient_info = parse_and_clean_data(file)
        if clean_df is None or clean_df.empty: print(f"Salto del file {file}."); continue
        
        image_paths = generate_static_visualizations(clean_df, output_folder)
        simulation_df = prepare_simulation_data(clean_df)
        weekly_model_df = create_weekly_model(clean_df)
        suspension_df = analyze_suspensions(clean_df, simulation_df.copy())
        hypo_analysis_df = analyze_hypo_episodes(clean_df)
        analysis_df = analyze_patterns_and_suggest(clean_df)
        
        base_name = os.path.splitext(os.path.basename(file))[0]
        excel_output_file = os.path.join(output_folder, f"Report_Interattivo_{base_name}.xlsx")
        try:
            create_interactive_excel_report(patient_info, clean_df, simulation_df, weekly_model_df, analysis_df, hypo_analysis_df, suspension_df, image_paths, excel_output_file)
        except PermissionError: print(f"\nERRORE: Permesso negato. Chiudere il file Excel '{excel_output_file}' se è aperto e riprovare.")
        except Exception as e: print(f"\nERRORE INASPETTATO durante la creazione del file Excel: {e}")
        
        for path in image_paths.values():
            if os.path.exists(path): os.remove(path)

if __name__ == '__main__':
    main()


